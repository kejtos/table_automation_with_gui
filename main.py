import os
from dataclasses import dataclass

from pandas import DataFrame, ExcelWriter
from pandas import to_datetime, wide_to_long, read_csv
from numpy import arange, append

from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

from unidecode import unidecode as ud
from datetime import datetime as dt


def import_tables(matice_vyroby_path:str, vyroba_path:str, komponenty_path:str, zavody_path:str):
    matice_vyroby = read_csv(matice_vyroby_path, sep="\t")
    komponenty = read_csv(komponenty_path, sep=";")
    vyroba = read_csv(vyroba_path, sep=";")
    zavody = read_csv(zavody_path, sep=";")
    dfs = (matice_vyroby, komponenty, vyroba, zavody)
    for df in dfs:
        df.columns = [ud(col) for col in df.columns]
    zavody['Místo'] = [ud(zavod) for zavod in zavody['Misto']]
    return (matice_vyroby, komponenty, vyroba, zavody)


def create_table(matice_vyroby:DataFrame, vyroba:DataFrame, komponenty:DataFrame, zavody:DataFrame, vystup:str):
    matice_vyroby = matice_vyroby.merge(komponenty, on='ID_komponenty', how='left')
    matice_vyroby['Cena_za_vstup'] = matice_vyroby['Mnozstvi']*matice_vyroby['Porizovaci_cena']
    ceny = matice_vyroby.groupby('ID_produktu').agg(Naklady_na_produkt=('Cena_za_vstup', sum)).reset_index()
    vyroba.loc[:,'Mesic'] = to_datetime(vyroba.loc[:, 'Datum'], format='%d.%m.%Y').dt.strftime('%Y/%m')
    produkt = vyroba.groupby(['ID_produktu', 'ID_zavodu', 'Mesic']).agg(Docasne_Mnozstvi=('Mnozstvi', sum)).reset_index()
    produkt = produkt.merge(ceny, on='ID_produktu', how='left')
    produkt['Docasne_Naklady'] = produkt['Docasne_Mnozstvi']*produkt['Naklady_na_produkt']
    produkt = produkt.merge(zavody, on='ID_zavodu', how='left')
    produkt = wide_to_long(produkt, stubnames='Docasne', i=['ID_zavodu', 'Místo', 'ID_produktu', 'Mesic'], j='Polozka',  sep='_', suffix=r'\w+').reset_index()
    df = produkt.pivot_table(index=['ID_zavodu', 'Misto', 'ID_produktu', 'Polozka'], columns='Mesic', values='Docasne', fill_value=0).reset_index()
    cols_name = [col.replace('/','_') for col in df.columns if '/' in col]
    time_now = dt.now().strftime('%Y_%m_%d %H-%M-%S')
    output_path = fr'{vystup}\Prehled {cols_name[-1]} - {time_now}.xlsx'

    b_rows = df['ID_zavodu'].duplicated('last') == False
    b_rows = df.index[b_rows].to_numpy()
    t_rows = append(0, b_rows[:-1]+1)

    # borders
    border_style = Side(border_style='medium', color='000000')
    l_border = Border(left=border_style)
    r_border = Border(right=border_style)
    t_border = Border(top=border_style)
    b_border = Border(bottom=border_style)
    tl_border = Border(left=border_style, top=border_style)
    bl_border = Border(left=border_style, bottom=border_style)
    tr_border = Border(right=border_style, top=border_style)
    br_border = Border(right=border_style, bottom=border_style)
    tb_border = Border(top=border_style, bottom=border_style)
    trb_border = Border(top=border_style, bottom=border_style, right=border_style)
    tlb_border = Border(top=border_style, bottom=border_style, left=border_style)
    borders = (l_border, r_border, t_border, b_border, tb_border, tl_border, tr_border, bl_border, br_border, trb_border, tlb_border)

    # fills
    exdark = PatternFill('solid', fgColor='95B3D7')
    dark = PatternFill('solid', fgColor='B8CCE4')
    light = PatternFill('solid', fgColor='DCE6F1')

    with ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name= 'Output', startcol=1, startrow=1, header=True, index=False)
        ws = writer.sheets['Output']

        f_row, l_row, f_col_n, l_col_n = ws.min_row, ws.max_row, ws.min_column, ws.max_column
        f_col, l_col = get_column_letter(f_col_n), get_column_letter(l_col_n)

        indeces = ws[f'{f_col}{f_row+1}':f'{get_column_letter(f_col_n+1)}{l_row}']
        for row in indeces:
            for cell in row:
                cell.fill = exdark

        headers = ws[f'{get_column_letter(f_col_n)}{f_row}':f'{l_col}{f_row}'][0]
        for cell in headers:
            cell.fill = exdark

        inside = ws[f'{get_column_letter(f_col_n+2)}{f_row+1}':f'{l_col}{l_row}']
        lenos = l_row-f_row
        light_rows = (arange(lenos) % 4 == 0) | (arange(lenos) % 4 == 1)

        for i, row in enumerate(inside):
            if light_rows[i]:
                for cell in row:
                    cell.fill = light
                    cell.number_format = '#,##0'
            else:
                for cell in row:
                    cell.fill = dark
                    cell.number_format = '#,##0'

        l_cells = [row[0] for row in indeces[1:-1]] + [row[0] for row in inside[1:-1]]
        r_cells = [row[-1] for row in indeces[1:-1]] + [row[-1] for row in inside[1:-1]]
        t_cells = [cell for i in t_rows for cell in inside[i][1:-1]]
        b_cells = [cell for i in b_rows for cell in inside[i][1:-1]]
        tb_cells = headers[2:-1]
        tl_cells = [indeces[i][0] for i in t_rows] + [inside[i][0] for i in t_rows]
        tr_cells = [indeces[i][-1] for i in t_rows] + [inside[i][-1] for i in t_rows]
        bl_cells = [indeces[i][0] for i in b_rows] + [inside[i][0] for i in b_rows]
        br_cells = [indeces[i][-1] for i in b_rows] + [inside[i][-1] for i in b_rows]
        tlb_cells = headers[0], headers[2]
        trb_cells = headers[1], headers[-1]

        ranges = (l_cells, r_cells, t_cells, b_cells, tb_cells, tl_cells, tr_cells, bl_cells, br_cells, trb_cells, tlb_cells)
        for rn, bor in zip(ranges, borders):
            for cell in rn:
                cell.border = bor

        for col in range(f_col_n, l_col_n+2):
            ws.column_dimensions[get_column_letter(col)].width = 12

        for i, val in enumerate(b_rows, start=1):
            ws.insert_rows(1 + val + f_row + i)

        for col in (f_col, get_column_letter(f_col_n+1)):
            for i, b_row in enumerate(b_rows):
                cell = ws[f'{col}{t_rows[i]+3+i}']
                ws.merge_cells(f'{col}{t_rows[i]+3+i}:{col}{b_row+3+i}')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)

        ws.insert_cols(f_col_n+2)
        ws.column_dimensions[get_column_letter(f_col_n+2)].width = 1


def main():
    parent_path = os.getcwd()
    inputs_path = fr'{parent_path}\vstupy'
    input_folders = os.listdir(inputs_path)
    initialdir = fr'{parent_path}\vstupy\{max(input_folders)}'
    path_names = ('komponenty', 'matice_vyroby', 'vyroba', 'zavody', 'výstup')
    basenames = ('komponenty.csv', 'matice_vyroby.txt', 'vyroba.txt', 'zavody.csv')
    outputdir = fr'{parent_path}\prehledy'

    @dataclass
    class GlobalData:
        initialdir: str
        path_names: tuple[str]
        outputdir: str
        basenames: tuple[str]
        file_paths = None

        def __post_init__(self):
            if self.file_paths is None:
                self.file_paths = {key: fr'{initialdir}\{name}' for key, name in zip(self.path_names[0:-1], self.basenames)}

    d = GlobalData(initialdir=initialdir, path_names=path_names, outputdir=outputdir, basenames=basenames)

    matice_vyroby, komponenty, vyroba, zavody = import_tables(komponenty_path=d.file_paths['komponenty'], matice_vyroby_path=d.file_paths['matice_vyroby'], vyroba_path=d.file_paths['vyroba'], zavody_path=d.file_paths['zavody'])
    create_table(matice_vyroby=matice_vyroby, vyroba=vyroba, komponenty=komponenty, zavody=zavody, vystup=d.outputdir)


if __name__ == '__main__':
    main()